"""
KursKompass API – Flask Backend (für Render.com)
Pro-User Datenspeicherung: Jeder User hat eigene Tree/Veranstaltungen.
"""
import os
import re
import json
import threading
from datetime import datetime
from io import BytesIO

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from scraper import (
    QISScraper, tree_to_dict, dict_to_tree
)

app = Flask(__name__)
CORS(app)

# User-Verwaltung: "user1:pass1,user2:pass2"
USERS_STR = os.environ.get("USERS", "loni:kurskompass2026")
USERS = {}
for pair in USERS_STR.split(","):
    pair = pair.strip()
    if ":" in pair:
        u, p = pair.split(":", 1)
        USERS[u.strip()] = p.strip()

API_KEY = os.environ.get("API_KEY", "loni-kurskompass-2026-secret")

# Pro-User Caches: { "loni": {"tree": [...], "veranstaltungen": [...]} }
user_caches = {}
cached_lehramtstypen = None
scraper_lock = threading.Lock()
DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
os.makedirs(DATA_DIR, exist_ok=True)


def check_auth():
    key = request.headers.get("X-API-Key") or request.args.get("key")
    if key != API_KEY:
        return jsonify({"error": "Unauthorized"}), 401
    return None


def get_user():
    """Holt Username aus Header."""
    return request.headers.get("X-User") or request.args.get("user") or "default"


def get_user_cache(user):
    if user not in user_caches:
        user_caches[user] = {"tree": None, "veranstaltungen": None}
    return user_caches[user]


def user_file(user, name):
    """Gibt Dateipfad für user-spezifische Datei zurück."""
    return os.path.join(DATA_DIR, f"{name}_{user}.json")


@app.route("/")
def health():
    return jsonify({"status": "ok", "app": "KursKompass API", "version": "1.2"})


@app.route("/api/ping")
def ping():
    return jsonify({"status": "awake"})


@app.route("/api/login", methods=["POST"])
def api_login():
    data = request.json or {}
    username = data.get("username", "").strip()
    password = data.get("password", "").strip()

    if username in USERS and USERS[username] == password:
        return jsonify({"status": "ok", "user": username, "api_key": API_KEY})
    else:
        return jsonify({"status": "error", "message": "Falscher Benutzername oder Passwort"}), 401


@app.route("/api/lehramtstypen")
def api_lehramtstypen():
    auth_error = check_auth()
    if auth_error:
        return auth_error

    global cached_lehramtstypen

    lt_file = os.path.join(DATA_DIR, "lehramtstypen.json")
    if cached_lehramtstypen:
        return jsonify(cached_lehramtstypen)
    elif os.path.exists(lt_file):
        with open(lt_file, "r", encoding="utf-8") as f:
            cached_lehramtstypen = json.load(f)
            return jsonify(cached_lehramtstypen)

    scraper = QISScraper()
    typen = scraper.scan_top_level()
    if typen:
        cached_lehramtstypen = typen
        with open(lt_file, "w", encoding="utf-8") as f:
            json.dump(typen, f, ensure_ascii=False, indent=2)
    return jsonify(typen)


@app.route("/api/scan-tree", methods=["POST"])
def api_scan_tree():
    auth_error = check_auth()
    if auth_error:
        return auth_error

    user = get_user()

    if scraper_lock.locked():
        return jsonify({"error": "Scraping läuft bereits. Bitte warte bis der aktuelle Scan fertig ist."}), 409

    start_roots = None
    if request.json:
        # Mehrere Studiengänge oder einzelner
        start_roots = request.json.get("root_paths")
        if not start_roots:
            single = request.json.get("root_path")
            if single:
                start_roots = [single]

    def do_scan():
        cache = get_user_cache(user)
        with scraper_lock:
            scraper = QISScraper()
            app.config["current_scraper"] = scraper
            if start_roots and len(start_roots) > 0:
                # Mehrere Roots nacheinander scannen, Ergebnisse zusammenführen
                all_trees = []
                for root in start_roots:
                    partial = scraper.scan_tree(start_root=root)
                    all_trees.extend(partial)
                tree = all_trees
            else:
                tree = scraper.scan_tree()
            cache["tree"] = tree
            tree_data = tree_to_dict(tree)
            with open(user_file(user, "tree"), "w", encoding="utf-8") as f:
                json.dump(tree_data, f, ensure_ascii=False, indent=2)
            # Don't pop scraper - let polling read final "scan_done" state

    thread = threading.Thread(target=do_scan)
    thread.start()
    return jsonify({"status": "started"})


@app.route("/api/tree")
def api_get_tree():
    auth_error = check_auth()
    if auth_error:
        return auth_error

    user = get_user()
    cache = get_user_cache(user)

    tf = user_file(user, "tree")
    if cache["tree"]:
        return jsonify(tree_to_dict(cache["tree"]))
    elif os.path.exists(tf):
        with open(tf, "r", encoding="utf-8") as f:
            data = json.load(f)
            cache["tree"] = dict_to_tree(data)
            return jsonify(data)
    return jsonify(None)


@app.route("/api/scrape", methods=["POST"])
def api_scrape():
    auth_error = check_auth()
    if auth_error:
        return auth_error

    user = get_user()
    cache = get_user_cache(user)

    if scraper_lock.locked():
        return jsonify({"error": "Scraping läuft bereits. Bitte warte bis der aktuelle Vorgang fertig ist."}), 409

    selected = request.json.get("selected", [])
    if not selected:
        return jsonify({"error": "Keine Bereiche ausgewählt"}), 400

    if not cache["tree"]:
        tf = user_file(user, "tree")
        if os.path.exists(tf):
            with open(tf, "r", encoding="utf-8") as f:
                cache["tree"] = dict_to_tree(json.load(f))
        else:
            return jsonify({"error": "Bitte zuerst Struktur laden"}), 400

    def do_scrape():
        with scraper_lock:
            scraper = QISScraper()
            app.config["current_scraper"] = scraper
            veranstaltungen = scraper.scrape_selected(cache["tree"], set(selected))
            from dataclasses import asdict
            ver_data = [asdict(v) for v in veranstaltungen]
            cache["veranstaltungen"] = ver_data
            with open(user_file(user, "veranstaltungen"), "w", encoding="utf-8") as f:
                json.dump(ver_data, f, ensure_ascii=False, indent=2)
            # Don't pop scraper - let polling read final "done" state

    thread = threading.Thread(target=do_scrape)
    thread.start()
    return jsonify({"status": "started"})


@app.route("/api/progress")
def api_progress():
    auth_error = check_auth()
    if auth_error:
        return auth_error

    scraper = app.config.get("current_scraper")
    if scraper:
        return jsonify(scraper.progress)
    return jsonify({"phase": "idle", "status": "Bereit", "current": 0, "total": 0, "details": []})


@app.route("/api/veranstaltungen")
def api_veranstaltungen():
    auth_error = check_auth()
    if auth_error:
        return auth_error

    user = get_user()
    cache = get_user_cache(user)

    vf = user_file(user, "veranstaltungen")
    if cache["veranstaltungen"]:
        return jsonify({"data": cache["veranstaltungen"], "count": len(cache["veranstaltungen"])})
    elif os.path.exists(vf):
        with open(vf, "r", encoding="utf-8") as f:
            cache["veranstaltungen"] = json.load(f)
            return jsonify({"data": cache["veranstaltungen"], "count": len(cache["veranstaltungen"])})
    return jsonify({"data": None, "count": 0})


@app.route("/api/download-excel")
def api_download_excel():
    auth_error = check_auth()
    if auth_error:
        return auth_error

    user = get_user()
    cache = get_user_cache(user)

    if not cache["veranstaltungen"]:
        vf = user_file(user, "veranstaltungen")
        if os.path.exists(vf):
            with open(vf, "r", encoding="utf-8") as f:
                cache["veranstaltungen"] = json.load(f)

    if not cache["veranstaltungen"]:
        return jsonify({"error": "Keine Daten vorhanden"}), 404

    ver_data = cache["veranstaltungen"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Stundenplan"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="D4467E", end_color="D4467E", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = [
        "Titel", "Bereich", "Modul", "Art", "Gruppe", "Dozent",
        "Tag", "Zeit", "Rhythmus", "Gebäude", "Raum", "SWS",
        "Max. TN", "Belegung", "Semester", "Studiengänge"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, v in enumerate(ver_data, 2):
        raum = v.get("raum", "")
        # Gebäude aus Raum extrahieren
        geb_match = re.match(r'^(.+?)\s*\d', raum)
        gebaeude = geb_match.group(1).rstrip(" -") if geb_match else raum.split(" - ")[0] if raum else ""

        values = [
            v.get("titel", ""), v.get("pfad", ""), v.get("kennung", ""),
            v.get("veranstaltungsart", ""), v.get("gruppe", ""), v.get("dozent", ""),
            v.get("tag", ""), v.get("zeit", ""), v.get("rhythmus", ""),
            gebaeude, raum, v.get("sws", ""), v.get("max_teilnehmer", ""),
            v.get("belegung", ""), v.get("semester", ""), v.get("studiengaenge", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    widths = [40, 30, 12, 20, 12, 25, 6, 16, 8, 25, 30, 5, 8, 12, 12, 40]
    for i, w in enumerate(widths):
        col_letter = chr(65 + i)
        ws.column_dimensions[col_letter].width = w

    ws.auto_filter.ref = f"A1:P{len(ver_data) + 1}"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"KursKompass_{user}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
